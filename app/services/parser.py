import xlrd
from datetime import datetime
from typing import List, Dict
import io

def parse_excel(file_bytes: bytes, filename: str = None) -> List[Dict]:
    """Парсер для Excel: поддерживает .xlsx (openpyxl) и .xls (xlrd)"""
    
    print(f"=== НАЧАЛО ПАРСИНГА ===")
    print(f"Имя файла: {filename}")
    
    is_xls = filename and filename.lower().endswith('.xls')
    transactions = []
    
    try:
        if is_xls:
            # Для .xls используем xlrd
            print("Использую xlrd для .xls файла")
            workbook = xlrd.open_workbook(file_contents=file_bytes)
            sheet = workbook.sheet_by_index(0)
            
            print(f"Строк: {sheet.nrows}, колонок: {sheet.ncols}")
            
            # Заголовки находятся в строке 1 (индекс 1), а не в 0
            headers = []
            for col_idx in range(sheet.ncols):
                cell_value = sheet.cell_value(1, col_idx)
                if cell_value:
                    cell_str = str(cell_value).strip().upper()
                    if cell_str and cell_str != "ВСЕГО":
                        # ИСКЛЮЧАЕМ КАТЕГОРИЮ "РАБОТА"
                        category_name = str(cell_value).strip().lower()
                        if category_name != "работа":
                            headers.append({
                                "col": col_idx,
                                "name": category_name
                            })
            
            print(f"Найдено категорий: {len(headers)}")
            if headers:
                print(f"Первые 5 категорий: {[h['name'] for h in headers[:5]]}")
            
            # Читаем данные (начиная со строки 2, индекс 2)
            for row_idx in range(2, sheet.nrows):
                month_cell = sheet.cell_value(row_idx, 0)
                if not month_cell:
                    continue
                
                month_str = str(month_cell).strip().upper()
                print(f"Обработка строки {row_idx}: месяц = {month_str}")
                
                if month_str == "ВСЕГО":
                    print("Достигнута строка ВСЕГО, остановка")
                    break
                
                # Преобразуем месяц в число
                month_map = {
                    'январь': 1, 'января': 1, 'янв': 1,
                    'февраль': 2, 'февраля': 2, 'фев': 2,
                    'март': 3, 'марта': 3, 'мар': 3,
                    'апрель': 4, 'апреля': 4, 'апр': 4,
                    'май': 5, 'мая': 5,
                    'июнь': 6, 'июня': 6, 'июн': 6,
                    'июль': 7, 'июля': 7, 'июл': 7,
                    'август': 8, 'августа': 8, 'авг': 8,
                    'сентябрь': 9, 'сентября': 9, 'сен': 9, 'сент': 9,
                    'октябрь': 10, 'октября': 10, 'окт': 10,
                    'ноябрь': 11, 'ноября': 11, 'ноя': 11,
                    'декабрь': 12, 'декабря': 12, 'дек': 12,
                }
                month_num = month_map.get(str(month_cell).lower(), 6)
                date_val = datetime(2026, month_num, 15)
                
                for header in headers:
                    # ИСКЛЮЧАЕМ КАТЕГОРИЮ "РАБОТА" (ещё раз для надёжности)
                    if header["name"] == "работа":
                        continue
                    
                    amount = sheet.cell_value(row_idx, header["col"])
                    
                    if amount is None or amount == "":
                        continue
                    
                    if isinstance(amount, str):
                        continue
                    
                    if isinstance(amount, (int, float)) and amount != 0:
                        transactions.append({
                            "date": date_val,
                            "amount": abs(float(amount)),
                            "category": header["name"],
                            "description": f"{month_cell} - {header['name']}: {amount} ₽"
                        })
            
            print(f"Всего транзакций: {len(transactions)}")
            return transactions
            
        else:
            # Для .xlsx используем openpyxl
            import openpyxl
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = workbook.active
            
            # Заголовки в строке 2 (индекс 2)
            headers = []
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=2, column=col_idx)
                if cell.value:
                    cell_str = str(cell.value).strip().upper()
                    if cell_str and cell_str != "ВСЕГО":
                        category_name = str(cell.value).strip().lower()
                        if category_name != "работа":
                            headers.append({
                                "col": col_idx,
                                "name": category_name
                            })
            
            for row_idx in range(3, sheet.max_row + 1):
                month_cell = sheet.cell(row=row_idx, column=1)
                month_name = month_cell.value
                
                if not month_name:
                    continue
                
                month_str = str(month_name).strip().upper()
                
                if month_str == "ВСЕГО":
                    break
                
                month_map = {
                    'январь': 1, 'января': 1, 'янв': 1,
                    'февраль': 2, 'февраля': 2, 'фев': 2,
                    'март': 3, 'марта': 3, 'мар': 3,
                    'апрель': 4, 'апреля': 4, 'апр': 4,
                    'май': 5, 'мая': 5,
                    'июнь': 6, 'июня': 6, 'июн': 6,
                    'июль': 7, 'июля': 7, 'июл': 7,
                    'август': 8, 'августа': 8, 'авг': 8,
                    'сентябрь': 9, 'сентября': 9, 'сен': 9, 'сент': 9,
                    'октябрь': 10, 'октября': 10, 'окт': 10,
                    'ноябрь': 11, 'ноября': 11, 'ноя': 11,
                    'декабрь': 12, 'декабря': 12, 'дек': 12,
                }
                month_num = month_map.get(str(month_name).lower(), 6)
                date_val = datetime(2026, month_num, 15)
                
                for header in headers:
                    if header["name"] == "работа":
                        continue
                    
                    amount_cell = sheet.cell(row=row_idx, column=header["col"])
                    amount = amount_cell.value
                    
                    if amount and isinstance(amount, (int, float)) and amount != 0:
                        transactions.append({
                            "date": date_val,
                            "amount": abs(float(amount)),
                            "category": header["name"],
                            "description": f"{month_name} - {header['name']}: {amount} ₽"
                        })
            
            print(f"Всего транзакций: {len(transactions)}")
            return transactions
            
    except Exception as e:
        print(f"ОШИБКА ПАРСИНГА: {e}")
        import traceback
        traceback.print_exc()
        return []
